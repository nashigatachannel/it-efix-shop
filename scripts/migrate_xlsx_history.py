"""xlsxの「シート1」履歴データを新Sheets「過去注文」タブにコピー移行する。

- xlsx: .tmp_efix_sales.xlsx 「シート1」(行6=ヘッダー、行7〜=データ)
- 移行先: 新Sheets「過去注文」タブ
- 通し番号空間はWeb注文と別系統 (xlsxは1,2,3...で手動採番)
- 空行はスキップ (通し番号セルが空 or A列が空)
"""

import sys
sys.stdout.reconfigure(encoding="utf-8")

from datetime import date, datetime

from google.oauth2 import service_account
from googleapiclient.discovery import build
from openpyxl import load_workbook

SA_PATH = r"C:\Users\tunag\accounting\credentials\efix-shop-service-account.json"
SHEET_ID = "1QNUcpyDnXc12GU8Uvxm_tMX1ZGG09Sm4azIKFAPrjTY"
XLSX_PATH = r"C:\Users\tunag\it-efix-shop\.tmp_efix_sales.xlsx"
DEST_TAB = "過去注文"
HEADER_ROW = 6
USED_COLS = 22  # A〜V列のみ (それ以降は集計エリア)


def cell_to_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    return str(v)


def main() -> None:
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["シート1"]

    # 1) ヘッダー抽出
    headers = []
    for ci in range(1, USED_COLS + 1):
        v = ws.cell(row=HEADER_ROW, column=ci).value
        headers.append(cell_to_str(v) if v else f"col{ci}")
    print(f"[HEADER] {len(headers)} cols: {headers[:6]}...")

    # 2) データ抽出 (通し番号セルが数値・文字列で値を持つ行のみ)
    rows = []
    for ri in range(HEADER_ROW + 1, ws.max_row + 1):
        first = ws.cell(row=ri, column=1).value
        if first is None or (isinstance(first, str) and not first.strip()):
            continue
        row = [cell_to_str(ws.cell(row=ri, column=ci).value) for ci in range(1, USED_COLS + 1)]
        rows.append(row)
    print(f"[DATA] {len(rows)} valid rows extracted")

    # 3) Sheets API: 「過去注文」タブを追加 or 既存ならクリア
    creds = service_account.Credentials.from_service_account_file(
        SA_PATH, scopes=["https://www.googleapis.com/auth/spreadsheets"]
    )
    sheets = build("sheets", "v4", credentials=creds, cache_discovery=False)

    meta = sheets.spreadsheets().get(
        spreadsheetId=SHEET_ID,
        fields="sheets(properties(sheetId,title))",
    ).execute()
    existing = {s["properties"]["title"]: s["properties"]["sheetId"] for s in meta["sheets"]}

    requests = []
    if DEST_TAB in existing:
        # クリアのみ
        sheets.spreadsheets().values().clear(
            spreadsheetId=SHEET_ID,
            range=f"{DEST_TAB}!A:Z",
        ).execute()
        # サイズ調整
        requests.append({
            "updateSheetProperties": {
                "properties": {
                    "sheetId": existing[DEST_TAB],
                    "gridProperties": {
                        "rowCount": max(1000, len(rows) + 10),
                        "columnCount": USED_COLS,
                        "frozenRowCount": 1,
                    },
                },
                "fields": "gridProperties.rowCount,gridProperties.columnCount,gridProperties.frozenRowCount",
            }
        })
    else:
        requests.append({
            "addSheet": {
                "properties": {
                    "title": DEST_TAB,
                    "gridProperties": {
                        "rowCount": max(1000, len(rows) + 10),
                        "columnCount": USED_COLS,
                        "frozenRowCount": 1,
                    },
                }
            }
        })

    if requests:
        sheets.spreadsheets().batchUpdate(
            spreadsheetId=SHEET_ID,
            body={"requests": requests},
        ).execute()
        print(f"[OK] tab '{DEST_TAB}' prepared")

    # 4) ヘッダー + データを一括書き込み
    values = [headers] + rows
    sheets.spreadsheets().values().update(
        spreadsheetId=SHEET_ID,
        range=f"{DEST_TAB}!A1",
        valueInputOption="USER_ENTERED",
        body={"values": values},
    ).execute()
    print(f"[OK] wrote {len(values)} rows × {USED_COLS} cols")

    # 5) ヘッダー書式
    sheet_id_target = existing.get(DEST_TAB)
    if sheet_id_target is None:
        # addSheet後の新IDを再取得
        meta2 = sheets.spreadsheets().get(
            spreadsheetId=SHEET_ID,
            fields="sheets(properties(sheetId,title))",
        ).execute()
        for s in meta2["sheets"]:
            if s["properties"]["title"] == DEST_TAB:
                sheet_id_target = s["properties"]["sheetId"]
                break

    sheets.spreadsheets().batchUpdate(
        spreadsheetId=SHEET_ID,
        body={
            "requests": [
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id_target,
                            "startRowIndex": 0,
                            "endRowIndex": 1,
                            "startColumnIndex": 0,
                            "endColumnIndex": USED_COLS,
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.20, "green": 0.20, "blue": 0.30},
                                "textFormat": {
                                    "foregroundColor": {"red": 1, "green": 1, "blue": 1},
                                    "bold": True,
                                },
                                "horizontalAlignment": "CENTER",
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)",
                    }
                }
            ]
        },
    ).execute()
    print("[OK] header formatted")
    print(f"\nDONE. https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit#gid=...{DEST_TAB}")


if __name__ == "__main__":
    main()
