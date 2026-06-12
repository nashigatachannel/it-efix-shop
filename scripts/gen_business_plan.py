"""
gen_business_plan.py
東京海上日動 請負業者賠償責任保険 引受審査用 事業計画書 生成スクリプト
出力: c:/Users/tunag/it-efix-shop/事業計画書_IT_2026.xlsx
"""

from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 定数
# ---------------------------------------------------------------------------
OUT_PATH = Path("c:/Users/tunag/it-efix-shop/事業計画書_IT_2026.xlsx")

NAVY = "1F3864"
LIGHT_BLUE = "D6E4F0"
LIGHT_YELLOW = "FFFACD"
GRAY = "F2F2F2"

FMT_YEN = '"¥"#,##0'
FMT_PCT = "0.0%"

FONT_BODY = "Yu Gothic"
FONT_SIZE_BODY = 10
FONT_SIZE_HEADER = 12
FONT_SIZE_TITLE = 24
FONT_SIZE_SUB = 14

TAB_COLOR = "1F3864"

THIN = Side(style="thin")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_BOTTOM = Border(bottom=THIN)


# ---------------------------------------------------------------------------
# ヘルパー
# ---------------------------------------------------------------------------
def navy_font(size=FONT_SIZE_BODY, bold=False):
    return Font(name=FONT_BODY, size=size, bold=bold, color="FFFFFF")


def body_font(size=FONT_SIZE_BODY, bold=False, color="000000"):
    return Font(name=FONT_BODY, size=size, bold=bold, color=color)


def navy_fill():
    return PatternFill("solid", fgColor=NAVY)


def light_fill(color=LIGHT_BLUE):
    return PatternFill("solid", fgColor=color)


def center_align(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def left_align(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def right_align():
    return Alignment(horizontal="right", vertical="center")


def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height


def apply_header_row(ws, row_idx, values, col_start=1):
    """紺背景・白太字の見出し行"""
    for i, val in enumerate(values):
        col = col_start + i
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font = navy_font(size=FONT_SIZE_HEADER, bold=True)
        cell.fill = navy_fill()
        cell.alignment = center_align(wrap=True)
        cell.border = BORDER_ALL


def apply_data_row(ws, row_idx, values, formats=None, col_start=1,
                   fill=None, bold=False):
    """データ行（罫線あり）"""
    if formats is None:
        formats = [None] * len(values)
    for i, (val, fmt) in enumerate(zip(values, formats)):
        col = col_start + i
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font = body_font(size=FONT_SIZE_BODY, bold=bold)
        cell.border = BORDER_ALL
        cell.alignment = left_align()
        if fmt:
            cell.number_format = fmt
        if val is not None and isinstance(val, (int, float)) and not isinstance(val, bool):
            cell.alignment = right_align()
        if fill:
            cell.fill = fill


def set_print_setup(ws):
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.tabColor = TAB_COLOR
    ws.page_margins.left = 0.7 / 2.54
    ws.page_margins.right = 0.7 / 2.54
    ws.page_margins.top = 1.0 / 2.54
    ws.page_margins.bottom = 1.0 / 2.54


# ---------------------------------------------------------------------------
# Sheet1: 表紙
# ---------------------------------------------------------------------------
def make_cover(wb):
    ws = wb.create_sheet("表紙")
    set_col_widths(ws, {"A": 4, "B": 22, "C": 14, "D": 14, "E": 14, "F": 14})
    set_print_setup(ws)

    # タイトル
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "事業計画書"
    c.font = Font(name=FONT_BODY, size=FONT_SIZE_TITLE, bold=True, color=NAVY)
    c.alignment = center_align()
    set_row_height(ws, 1, 60)

    # サブタイトル
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = "請負業者賠償責任保険 加入申込資料"
    c.font = Font(name=FONT_BODY, size=FONT_SIZE_SUB, bold=True, color=NAVY)
    c.alignment = center_align()
    set_row_height(ws, 2, 36)

    ws.merge_cells("A3:F3")
    set_row_height(ws, 3, 12)

    # 区切り線
    for col in range(1, 7):
        cell = ws.cell(row=4, column=col)
        cell.border = Border(bottom=Side(style="medium", color=NAVY))
    set_row_height(ws, 4, 4)

    # 事業者情報
    info_rows = [
        ("屋号・名称", "IT"),
        ("代表者氏名", "石川 卓磨"),
        ("所在地", "〒062-0041 北海道札幌市豊平区福住一条7丁目4-13"),
        ("電話番号", "080-6282-4834"),
        ("メールアドレス", "takuma.ishikawa.line@gmail.com"),
        ("公式サイト", "https://it-efix-shop.vercel.app"),
        ("作成日", "2026年5月12日"),
    ]
    start_row = 5
    for i, (label, val) in enumerate(info_rows):
        r = start_row + i
        ws.merge_cells(f"A{r}:B{r}")
        ws.merge_cells(f"C{r}:F{r}")
        lc = ws[f"A{r}"]
        lc.value = label
        lc.font = body_font(bold=True)
        lc.alignment = left_align()
        lc.fill = light_fill(GRAY)
        lc.border = BORDER_ALL
        vc = ws[f"C{r}"]
        vc.value = val
        vc.font = body_font()
        vc.alignment = left_align()
        vc.border = BORDER_ALL
        set_row_height(ws, r, 20)

    # 提出先
    sep_row = start_row + len(info_rows) + 1
    ws.merge_cells(f"A{sep_row}:F{sep_row}")
    set_row_height(ws, sep_row, 30)

    dest_row = sep_row + 1
    ws.merge_cells(f"A{dest_row}:F{dest_row}")
    c = ws[f"A{dest_row}"]
    c.value = "提出先: 東京海上日動パートナーズ北海道札幌支社 御中"
    c.font = body_font(size=12, bold=True, color=NAVY)
    c.alignment = center_align()
    set_row_height(ws, dest_row, 28)

    sub_row = dest_row + 1
    ws.merge_cells(f"A{sub_row}:F{sub_row}")
    c = ws[f"A{sub_row}"]
    c.value = "営業担当: 若松 章愛 様"
    c.font = body_font(size=11)
    c.alignment = center_align()
    set_row_height(ws, sub_row, 24)


# ---------------------------------------------------------------------------
# Sheet2: 事業者概要
# ---------------------------------------------------------------------------
def make_overview(wb):
    ws = wb.create_sheet("事業者概要")
    set_col_widths(ws, {"A": 4, "B": 28, "C": 30, "D": 14, "E": 14, "F": 14})
    set_print_setup(ws)

    # タイトル
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "事業者概要"
    c.font = Font(name=FONT_BODY, size=FONT_SIZE_HEADER, bold=True, color="FFFFFF")
    c.fill = navy_fill()
    c.alignment = center_align()
    set_row_height(ws, 1, 30)

    rows = [
        ("屋号", "IT"),
        ("代表者氏名", "石川 卓磨"),
        ("生年月日", "1995年11月14日"),
        ("開業年月日", "2026年3月1日"),
        ("事業形態", "個人事業主"),
        ("所在地", "〒062-0041 北海道札幌市豊平区福住一条7丁目4-13"),
        ("電話番号", "080-6282-4834"),
        ("メールアドレス", "takuma.ishikawa.line@gmail.com"),
        ("インボイス登録番号", "T2810703528253（令和8年1月1日、札幌南税務署）"),
        ("取引銀行①", "楽天銀行（個人口座）"),
        ("取引銀行②", "GMOあおぞらネット銀行（個人事業主口座、開設手続中）"),
        ("公式サイト", "https://it-efix-shop.vercel.app"),
    ]

    for i, (label, val) in enumerate(rows):
        r = i + 2
        ws.merge_cells(f"B{r}:B{r}")
        ws.merge_cells(f"C{r}:F{r}")
        lc = ws.cell(row=r, column=2, value=label)
        lc.font = body_font(bold=True)
        lc.alignment = left_align()
        lc.fill = light_fill(GRAY)
        lc.border = BORDER_ALL
        vc = ws.cell(row=r, column=3, value=val)
        vc.font = body_font()
        vc.alignment = left_align(wrap=True)
        vc.border = BORDER_ALL
        # merge C to F
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        set_row_height(ws, r, 20)


# ---------------------------------------------------------------------------
# Sheet3: 事業内容
# ---------------------------------------------------------------------------
def make_business(wb):
    ws = wb.create_sheet("事業内容")
    set_col_widths(ws, {"A": 4, "B": 22, "C": 14, "D": 14, "E": 14, "F": 24})
    set_print_setup(ws)

    r = 1
    # タイトル
    ws.merge_cells(f"A{r}:F{r}")
    c = ws[f"A{r}"]
    c.value = "事業内容"
    c.font = navy_font(size=FONT_SIZE_HEADER, bold=True)
    c.fill = navy_fill()
    c.alignment = center_align()
    set_row_height(ws, r, 30)
    r += 1

    # 主要事業説明
    ws.merge_cells(f"A{r}:F{r}")
    c = ws[f"A{r}"]
    c.value = "■ 主要事業"
    c.font = body_font(bold=True, color=NAVY)
    c.alignment = left_align()
    set_row_height(ws, r, 20)
    r += 1

    ws.merge_cells(f"A{r}:F{r}")
    c = ws[f"A{r}"]
    c.value = "農業機械向け電動自動操舵装置「e-steer」（EFIX社製）の販売および取付業務"
    c.font = body_font()
    c.alignment = left_align(wrap=True)
    set_row_height(ws, r, 20)
    r += 2

    # 取扱商品テーブル
    ws.merge_cells(f"A{r}:F{r}")
    c = ws[f"A{r}"]
    c.value = "■ 取扱商品ラインナップ（税抜販売価格）"
    c.font = body_font(bold=True, color=NAVY)
    c.alignment = left_align()
    set_row_height(ws, r, 20)
    r += 1

    headers = ["略称", "製品名", "販売価格(税抜)", "仕入価格", "粗利", "主要スペック"]
    apply_header_row(ws, r, headers, col_start=1)
    set_row_height(ws, r, 22)
    r += 1

    products = [
        ("e-steer 20", "eSteer 20セット", 1045455, 450000, 595455,
         "10.1インチLCD、GPS高精度±2.5cm、ISOBUS対応"),
        ("e-steer 20 MAX", "eSteer 20MAX セット", 1181819, 500000, 681819,
         "12.1インチLCD、外車対応、ISOBUS対応"),
        ("e-steer 10", "eSteer 10セット", 909091, 410000, 499091,
         "スタンダードモデル（2026年4月販売停止）"),
    ]
    fills = [None, light_fill(GRAY), None]
    for j, (abr, name, sell, buy, margin, spec) in enumerate(products):
        apply_data_row(ws, r, [abr, name, sell, buy, margin, spec],
                       formats=[None, None, FMT_YEN, FMT_YEN, FMT_YEN, None],
                       fill=fills[j % 2])
        set_row_height(ws, r, 22)
        r += 1
    r += 1

    # 付帯サービス
    ws.merge_cells(f"A{r}:F{r}")
    c = ws[f"A{r}"]
    c.value = "■ 付帯サービス"
    c.font = body_font(bold=True, color=NAVY)
    c.alignment = left_align()
    set_row_height(ws, r, 20)
    r += 1

    services = [
        "取付工賃: ¥100,000/件（税抜、出張取付）",
        "取外し工賃: ¥50,000/件（税抜）",
        "ブラケット・スリーブ等オプション: ¥3,200〜¥75,000",
        "ISOBUSケーブル: ¥115,000/本",
        "無線リモコン: ¥25,000",
        "EFIX位置補正サービス（PointSky）: 年¥39,600（初年度無料）",
        "VRS位置補正: 永年無料",
    ]
    for svc in services:
        ws.merge_cells(f"A{r}:F{r}")
        c = ws[f"A{r}"]
        c.value = "  ・" + svc
        c.font = body_font()
        c.alignment = left_align()
        c.border = Border(left=THIN, right=THIN, bottom=THIN)
        set_row_height(ws, r, 18)
        r += 1
    r += 1

    # 業務体制
    ws.merge_cells(f"A{r}:F{r}")
    c = ws[f"A{r}"]
    c.value = "■ 業務体制"
    c.font = body_font(bold=True, color=NAVY)
    c.alignment = left_align()
    set_row_height(ws, r, 20)
    r += 1

    taisei = [
        "代表（石川卓磨）: 営業、商品手配、顧客対応、納品立会",
        "業務委託先（前田氏ほか）: 取付作業（個別案件単位の委託契約、副業整備士含む）",
        "主要稼働エリア: 北海道全域（札幌・帯広・士幌・更別・池田・富良野・幕別・上士幌・島根ほか）",
    ]
    for t in taisei:
        ws.merge_cells(f"A{r}:F{r}")
        c = ws[f"A{r}"]
        c.value = "  ・" + t
        c.font = body_font()
        c.alignment = left_align(wrap=True)
        set_row_height(ws, r, 20)
        r += 1


# ---------------------------------------------------------------------------
# Sheet4: 販売実績
# ---------------------------------------------------------------------------
def make_sales(wb):
    ws = wb.create_sheet("販売実績")
    set_col_widths(ws, {"A": 18, "B": 12, "C": 20, "D": 32, "E": 16})
    set_print_setup(ws)

    r = 1
    ws.merge_cells(f"A{r}:E{r}")
    c = ws[f"A{r}"]
    c.value = "販売実績（2026年4月〜5月12日）"
    c.font = navy_font(size=FONT_SIZE_HEADER, bold=True)
    c.fill = navy_fill()
    c.alignment = center_align()
    set_row_height(ws, r, 30)
    r += 1

    headers = ["請求書番号", "発行日", "取引先", "件名", "金額(税込)"]
    apply_header_row(ws, r, headers, col_start=1)
    set_row_height(ws, r, 22)
    r += 1

    invoices = [
        ("INV-0000000031", "2026-04-28", "（株）奥原商会", "eSteer 20MAX セット", 1111550),
        ("INV-0000000032", "2026-04-28", "中田", "eSteer 10セット", 873400),
        ("INV-0000000033", "2026-04-28", "出口", "eSteer 10セット", 825000),
        ("INV-0000000034", "2026-04-28", "（株）奥原商会", "eSteer 20MAX セット", 1072500),
        ("INV-0000000035", "2026-04-28", "（株）奥原商会", "eSteer 20MAX セット", 1042800),
        ("INV-0000000036", "2026-04-28", "河口和吉", "取外し工賃", 235400),
        ("INV-0000000084", "2026-04-29", "高橋サービス", "eSteer 20セット", 715000),
        ("INV-0000000085", "2026-04-29", "美馬大介", "eSteer 20MAX セット", 1426500),
        ("INV-0000000086", "2026-04-29", "河野 辰徳", "eSteer 20セット", 1700001),
        ("INV-0000000087", "2026-04-29", "松本寿弘", "eSteer 20セット", 500000),
        ("INV-0000000088", "2026-04-29", "藤井信二", "商品", 286000),
        ("INV-0000000094", "2026-05-03", "Tアグリ", "NMEAケーブル", 38500),
        ("INV-0000000095", "2026-05-05", "宗原", "eSteer 20MAX セット", 1300000),
        ("INV-0000000096", "2026-05-05", "森田", "eSteer 20MAX セット", 1300000),
        ("INV-0000000097", "2026-05-08", "岡本達幸", "eSteer 20セット", 1000000),
        ("INV-0000000098", "2026-05-08", "吉田", "eSteer 20MAX セット", 1300000),
        ("INV-0000000099", "2026-05-08", "HMK", "eSteer 20MAX セット", 770000),
        ("INV-0000000100", "2026-05-08", "斎藤ゆきや", "eSteer 20セット", 1150001),
        ("INV-0000000101", "2026-05-08", "予備 吉田宏一", "eSteer 20セット", 1287501),
        ("INV-0000000102", "2026-05-08", "高橋サービス", "無線リモコン", 17875),
        ("INV-0000000103", "2026-05-08", "菊池", "Integrated Main Cable", 101640),
        ("INV-0000000104", "2026-05-08", "山地 永存", "無線リモコン", 17875),
        ("INV-0000000105", "2026-05-09", "岡本 達幸", "取外し工賃", 50001),
        ("INV-0000000108", "2026-05-10", "吉田 貴洋", "eSteer 20セット", 1205001),
        ("INV-0000000109", "2026-05-12", "（株）大井アグリサポート", "eSteer 20MAX セット", 4620000),
    ]

    for j, inv in enumerate(invoices):
        fill = light_fill(GRAY) if j % 2 == 1 else None
        apply_data_row(ws, r, list(inv),
                       formats=[None, None, None, None, FMT_YEN],
                       fill=fill)
        set_row_height(ws, r, 18)
        r += 1

    # 合計行
    total = 24046545
    total_fill = PatternFill("solid", fgColor="1F3864")
    for col in range(1, 6):
        cell = ws.cell(row=r, column=col)
        cell.fill = total_fill
        cell.font = Font(name=FONT_BODY, size=FONT_SIZE_BODY, bold=True, color="FFFFFF")
        cell.border = BORDER_ALL
        if col == 4:
            cell.value = "合計（税込）"
            cell.alignment = right_align()
        elif col == 5:
            cell.value = total
            cell.number_format = FMT_YEN
            cell.alignment = right_align()
    set_row_height(ws, r, 22)
    r += 2

    # サマリーボックス
    ws.merge_cells(f"A{r}:E{r}")
    c = ws[f"A{r}"]
    c.value = "■ 簡易サマリー"
    c.font = body_font(bold=True, color=NAVY)
    c.alignment = left_align()
    set_row_height(ws, r, 20)
    r += 1

    summary = [
        ("売上合計（税込）", 24046545, FMT_YEN),
        ("売上合計（税抜換算 ÷1.1）", 21860495, FMT_YEN),
        ("期間", "2026年4月28日 〜 5月12日（約2週間）", None),
        ("取扱件数", "25件", None),
        ("平均単価（税込）", 961861, FMT_YEN),
    ]
    for label, val, fmt in summary:
        ws.merge_cells(f"A{r}:B{r}")
        ws.merge_cells(f"C{r}:E{r}")
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = body_font(bold=True)
        lc.fill = light_fill(GRAY)
        lc.border = BORDER_ALL
        lc.alignment = left_align()
        vc = ws.cell(row=r, column=3, value=val)
        vc.font = body_font(bold=True)
        vc.border = BORDER_ALL
        if fmt:
            vc.number_format = fmt
            vc.alignment = right_align()
        else:
            vc.alignment = left_align()
        set_row_height(ws, r, 20)
        r += 1


# ---------------------------------------------------------------------------
# Sheet5: 年商計画
# ---------------------------------------------------------------------------
def make_annual_plan(wb):
    ws = wb.create_sheet("年商計画")
    set_col_widths(ws, {"A": 16, "B": 22, "C": 36, "D": 14, "E": 14, "F": 14})
    set_print_setup(ws)

    r = 1
    ws.merge_cells(f"A{r}:C{r}")
    c = ws[f"A{r}"]
    c.value = "年商計画（2026年4月〜2027年3月）"
    c.font = navy_font(size=FONT_SIZE_HEADER, bold=True)
    c.fill = navy_fill()
    c.alignment = center_align()
    set_row_height(ws, r, 30)
    r += 1

    headers = ["月", "売上見込(税抜)", "内訳・備考"]
    apply_header_row(ws, r, headers, col_start=1)
    set_row_height(ws, r, 22)
    r += 1

    monthly = [
        ("2026年4月", 9890000, "実績（4月計上分）"),
        ("2026年5月", 15000000, "実績ベース＋見込（農繁期ピーク）"),
        ("2026年6月", 12000000, "田植え後、追加導入需要"),
        ("2026年7月", 9000000, "中干し期、防除関連"),
        ("2026年8月", 7000000, "夏期需要"),
        ("2026年9月", 6000000, "収穫前点検需要"),
        ("2026年10月", 6000000, "秋収穫期"),
        ("2026年11月", 5000000, "収穫後点検"),
        ("2026年12月", 5000000, "年末・冬季導入相談"),
        ("2027年1月", 4000000, "閑散期"),
        ("2027年2月", 5000000, "春準備需要立ち上がり"),
        ("2027年3月", 16110000, "春作業前ラストスパート"),
    ]

    for j, (month, amount, note) in enumerate(monthly):
        fill = light_fill(GRAY) if j % 2 == 1 else None
        apply_data_row(ws, r, [month, amount, note],
                       formats=[None, FMT_YEN, None],
                       fill=fill)
        set_row_height(ws, r, 18)
        r += 1

    # 合計行
    sum_fill = navy_fill()
    for col, (val, fmt) in enumerate(
        [("合計", None), (100000000, FMT_YEN), ("年商見込1億円", None)], start=1
    ):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font = Font(name=FONT_BODY, size=FONT_SIZE_BODY, bold=True, color="FFFFFF")
        cell.fill = sum_fill
        cell.border = BORDER_ALL
        if fmt:
            cell.number_format = fmt
            cell.alignment = right_align()
        else:
            cell.alignment = center_align() if col == 1 else left_align()
    set_row_height(ws, r, 22)
    r += 2

    # 商品別計画
    ws.merge_cells(f"A{r}:D{r}")
    c = ws[f"A{r}"]
    c.value = "■ 商品別販売計画"
    c.font = body_font(bold=True, color=NAVY)
    c.alignment = left_align()
    set_row_height(ws, r, 20)
    r += 1

    prod_headers = ["商品", "想定販売台数", "単価(税抜)", "売上小計"]
    apply_header_row(ws, r, prod_headers, col_start=1)
    set_row_height(ws, r, 22)
    r += 1

    prod_data = [
        ("e-steer 20", "50台", 1045455, 52272750),
        ("e-steer 20 MAX", "30台", 1181819, 35454570),
        ("オプション・取付工賃・部品", "一式", None, 12272680),
    ]
    for j, (name, qty, unit, subtotal) in enumerate(prod_data):
        fill = light_fill(GRAY) if j % 2 == 1 else None
        apply_data_row(ws, r, [name, qty, unit, subtotal],
                       formats=[None, None, FMT_YEN, FMT_YEN],
                       fill=fill)
        set_row_height(ws, r, 18)
        r += 1

    # 合計
    for col, (val, fmt) in enumerate(
        [("合計", None), ("80台+α", None), (None, None), (100000000, FMT_YEN)], start=1
    ):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font = Font(name=FONT_BODY, size=FONT_SIZE_BODY, bold=True, color="FFFFFF")
        cell.fill = navy_fill()
        cell.border = BORDER_ALL
        if fmt:
            cell.number_format = fmt
            cell.alignment = right_align()
        else:
            cell.alignment = left_align()
    set_row_height(ws, r, 22)


# ---------------------------------------------------------------------------
# Sheet6: 損益計画
# ---------------------------------------------------------------------------
def make_pnl(wb):
    ws = wb.create_sheet("損益計画")
    set_col_widths(ws, {"A": 4, "B": 36, "C": 18, "D": 14, "E": 14, "F": 14})
    set_print_setup(ws)

    r = 1
    ws.merge_cells(f"A{r}:D{r}")
    c = ws[f"A{r}"]
    c.value = "損益計画"
    c.font = navy_font(size=FONT_SIZE_HEADER, bold=True)
    c.fill = navy_fill()
    c.alignment = center_align()
    set_row_height(ws, r, 30)
    r += 1

    headers = ["項目", "金額", "構成比"]
    apply_header_row(ws, r, headers, col_start=2)
    set_row_height(ws, r, 22)
    r += 1

    pnl_rows = [
        ("売上高", 100000000, 1.0, False),
        ("売上原価（仕入）", 39500000, 0.395, False),
        ("　e-steer 20 仕入 (50台×¥450,000)", 22500000, None, False),
        ("　e-steer 20 MAX 仕入 (30台×¥500,000)", 15000000, None, False),
        ("　オプション・付属品仕入", 2000000, None, False),
        ("売上総利益（粗利）", 60500000, 0.605, True),
        ("販売費及び一般管理費", 16700000, 0.167, False),
        ("　取付外注費（業務委託）", 8000000, None, False),
        ("　旅費交通費（出張取付・営業）", 3000000, None, False),
        ("　通信費・システム費（EC・サーバ）", 1500000, None, False),
        ("　広告宣伝費（Web広告・展示会）", 2000000, None, False),
        ("　事務所・倉庫家賃光熱費", 1200000, None, False),
        ("　損害保険料（請負賠責・自動車保険等）", 500000, None, False),
        ("　雑費・消耗品費", 500000, None, False),
        ("営業利益", 43800000, 0.438, True),
    ]

    for label, amount, pct, is_total in pnl_rows:
        if is_total:
            fill = light_fill(LIGHT_BLUE)
            bold = True
        elif label.startswith("　"):
            fill = light_fill(GRAY)
            bold = False
        else:
            fill = None
            bold = False

        # col B: label
        lc = ws.cell(row=r, column=2, value=label)
        lc.font = body_font(bold=bold)
        lc.fill = fill if fill else PatternFill()
        lc.border = BORDER_ALL
        lc.alignment = left_align()

        # col C: amount
        ac = ws.cell(row=r, column=3, value=amount)
        ac.font = body_font(bold=bold)
        ac.fill = fill if fill else PatternFill()
        ac.border = BORDER_ALL
        ac.number_format = FMT_YEN
        ac.alignment = right_align()

        # col D: pct
        pc = ws.cell(row=r, column=4, value=pct)
        pc.font = body_font(bold=bold)
        pc.fill = fill if fill else PatternFill()
        pc.border = BORDER_ALL
        if pct is not None:
            pc.number_format = FMT_PCT
            pc.alignment = right_align()
        else:
            pc.value = "-"
            pc.alignment = center_align()

        set_row_height(ws, r, 20)
        r += 1


# ---------------------------------------------------------------------------
# Sheet7: 業務体制・リスクと保険
# ---------------------------------------------------------------------------
def make_risk(wb):
    ws = wb.create_sheet("業務体制・リスクと保険必要性")
    set_col_widths(ws, {"A": 4, "B": 22, "C": 34, "D": 30, "E": 14, "F": 14})
    set_print_setup(ws)

    r = 1
    ws.merge_cells(f"A{r}:D{r}")
    c = ws[f"A{r}"]
    c.value = "業務体制・想定リスクと保険必要性"
    c.font = navy_font(size=FONT_SIZE_HEADER, bold=True)
    c.fill = navy_fill()
    c.alignment = center_align()
    set_row_height(ws, r, 30)
    r += 1

    # 業務体制
    ws.merge_cells(f"A{r}:D{r}")
    c = ws[f"A{r}"]
    c.value = "■ 業務体制図"
    c.font = body_font(bold=True, color=NAVY)
    c.alignment = left_align()
    set_row_height(ws, r, 20)
    r += 1

    taisei = [
        "代表（石川卓磨）: 受注・販売・顧客対応・取付立会",
        "業務委託先：前田氏（主取付パートナー、副業整備士）ほか案件単位の外注",
        "取引先：エンドユーザー（個人農家）、卸先（武藤農機、奥原商会、大井アグリサポート、塚本氏、高橋サービス、山地氏ほか）",
    ]
    for t in taisei:
        ws.merge_cells(f"A{r}:D{r}")
        c = ws[f"A{r}"]
        c.value = "  ・" + t
        c.font = body_font()
        c.alignment = left_align(wrap=True)
        c.border = Border(left=THIN, right=THIN, bottom=THIN)
        set_row_height(ws, r, 22)
        r += 1
    r += 1

    # リスクテーブル
    ws.merge_cells(f"A{r}:D{r}")
    c = ws[f"A{r}"]
    c.value = "■ 想定リスクと補償必要性"
    c.font = body_font(bold=True, color=NAVY)
    c.alignment = left_align()
    set_row_height(ws, r, 20)
    r += 1

    risk_headers = ["リスク区分", "想定される事故事例", "補償ニーズ"]
    apply_header_row(ws, r, risk_headers, col_start=2)
    set_row_height(ws, r, 22)
    r += 1

    risks = [
        ("受託物損壊",
         "取付作業中の顧客トラクター本体損傷（操作ミス、配線ショート、誤接続）",
         "受託物賠償（顧客財物への賠償）"),
        ("第三者賠償",
         "取付作業中の第三者ケガ、第三者物損",
         "対人・対物賠償（生産物賠償）"),
        ("下請負人による事故",
         "業務委託先（前田氏ほか）の作業中事故",
         "下請負人特約"),
        ("施設管理",
         "事務所・倉庫での来訪者の事故、物損",
         "施設賠償"),
        ("完成作業危険",
         "取付後の不具合による顧客トラクター事故",
         "完成作業危険補償"),
    ]
    for j, (risk, example, need) in enumerate(risks):
        fill = light_fill(GRAY) if j % 2 == 1 else None
        apply_data_row(ws, r, [risk, example, need],
                       formats=[None, None, None],
                       fill=fill, col_start=2)
        set_row_height(ws, r, 28)
        r += 1
    r += 1

    # 加入希望補償
    ws.merge_cells(f"A{r}:D{r}")
    c = ws[f"A{r}"]
    c.value = "■ 加入希望補償"
    c.font = body_font(bold=True, color=NAVY)
    c.alignment = left_align()
    set_row_height(ws, r, 20)
    r += 1

    hope_items = [
        ("保険種目", "請負業者賠償責任保険"),
        ("補償限度額", "3,000万円"),
        ("契約方式", "年間包括契約方式"),
        ("特約", "下請負人特約・受託物損壊担保特約"),
        ("想定請負件数", "年間80件以上（実績ベース月平均6〜8件、繁忙期月15件）"),
    ]
    for label, val in hope_items:
        ws.merge_cells(f"A{r}:B{r}")
        ws.merge_cells(f"C{r}:D{r}")
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = body_font(bold=True)
        lc.fill = light_fill(GRAY)
        lc.border = BORDER_ALL
        lc.alignment = left_align()
        vc = ws.cell(row=r, column=3, value=val)
        vc.font = body_font()
        vc.border = BORDER_ALL
        vc.alignment = left_align(wrap=True)
        set_row_height(ws, r, 22)
        r += 1


# ---------------------------------------------------------------------------
# メイン
# ---------------------------------------------------------------------------
def main():
    wb = Workbook()
    # デフォルトシートを削除
    wb.remove(wb.active)

    make_cover(wb)
    make_overview(wb)
    make_business(wb)
    make_sales(wb)
    make_annual_plan(wb)
    make_pnl(wb)
    make_risk(wb)

    wb.save(OUT_PATH)
    print(f"保存完了: {OUT_PATH}")

    # 検証
    size = OUT_PATH.stat().st_size
    print(f"ファイルサイズ: {size:,} bytes ({size/1024:.1f} KB)")

    wb2 = load_workbook(OUT_PATH)
    print(f"\nシート一覧 ({len(wb2.sheetnames)} シート):")
    for name in wb2.sheetnames:
        ws = wb2[name]
        print(f"  [{name}]  {ws.max_row} 行 × {ws.max_column} 列")
    wb2.close()


if __name__ == "__main__":
    main()
