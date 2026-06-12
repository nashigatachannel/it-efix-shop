from __future__ import annotations

import json
import re
import shutil
import sys
from dataclasses import asdict, dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw


TAX_RATE = Decimal("0.10")
MAPPING_JSON = "EFIX_画像対応表.json"


@dataclass
class CatalogItem:
    id: str
    kind: str
    shortName: str
    model: str
    section: str
    category: str
    partNumber: str
    name: str
    requiredQty: int
    wholesalePriceExTax: int
    retailPriceExTax: int | None
    wholesalePriceIncTax: int
    retailPriceIncTax: int | None
    image: str
    sourceSheet: str
    sourceRow: int


def as_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value)).strip()


def as_int(value) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    except (InvalidOperation, ValueError):
        return 0


def price_inc_tax(price_ex_tax: int | None) -> int | None:
    if price_ex_tax is None:
        return None
    return int(
        (Decimal(price_ex_tax) * (Decimal("1") + TAX_RATE)).quantize(
            Decimal("1"),
            rounding=ROUND_HALF_UP,
        ),
    )


def slug(value: str, fallback: str) -> str:
    text = value.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = text.strip("-")
    return text or fallback


def normalize_part(value: str) -> str:
    text = as_text(value).upper()
    if text.endswith(".0"):
        text = text[:-2]
    text = re.sub(r"[^A-Z0-9]+", "", text)
    return text.lstrip("0") if text.isdigit() else text


def normalize_name(value: str) -> str:
    text = as_text(value).lower()
    text = text.replace("ｅ", "e").replace("　", " ")
    return re.sub(r"[^a-z0-9ぁ-んァ-ヶ一-龥]+", "", text)


def public_asset_name(index: int, suffix: str) -> str:
    return f"product-{index:03d}{suffix.lower() if suffix else '.png'}"


def create_placeholder(public_dir: Path) -> str:
    out = public_dir / "image-pending.png"
    image = Image.new("RGB", (640, 420), "#f4f4f2")
    draw = ImageDraw.Draw(image)
    draw.rounded_rectangle((110, 92, 530, 328), radius=12, outline="#d6d3ce", width=4)
    draw.line((170, 270, 280, 190, 365, 252, 455, 160), fill="#9ca3af", width=6)
    draw.ellipse((220, 135, 270, 185), fill="#d1d5db")
    image.save(out, "PNG", optimize=True)
    return "/wholesale-assets/image-pending.png"


def clean_generated_assets(public_dir: Path) -> None:
    public_dir.mkdir(parents=True, exist_ok=True)
    for pattern in ("product-*", "image-pending.png"):
        for path in public_dir.glob(pattern):
            if path.is_file():
                path.unlink()


def load_mapped_records(source_dir: Path) -> list[dict]:
    mapping_path = source_dir / MAPPING_JSON
    if not mapping_path.exists():
        return []
    records = json.loads(mapping_path.read_text(encoding="utf-8"))
    return [record for record in records if (source_dir / record.get("fileName", "")).exists()]


def copy_record_image(source: Path, out: Path) -> tuple[int, int]:
    with Image.open(source) as img:
        width, height = img.size
        if source.suffix.lower() == ".jp2":
            img.convert("RGB").save(out, "PNG", optimize=True)
        else:
            shutil.copy2(source, out)
    return width, height


def copy_images(source_dir: Path, public_dir: Path, generated_hero: Path) -> tuple[list[dict], str, str]:
    clean_generated_assets(public_dir)
    hero_out = public_dir / "wholesale-hero.png"
    shutil.copy2(generated_hero, hero_out)
    placeholder = create_placeholder(public_dir)

    assets: list[dict] = []
    mapped_records = load_mapped_records(source_dir)
    if mapped_records:
        for index, record in enumerate(mapped_records, 1):
            src = source_dir / record["fileName"]
            out_name = public_asset_name(index, ".png")
            out = public_dir / out_name
            width, height = copy_record_image(src, out)
            assets.append(
                {
                    "sourceName": record["fileName"],
                    "path": f"/wholesale-assets/{out_name}",
                    "width": width,
                    "height": height,
                    "sheet": record.get("sheet", ""),
                    "sourceRow": record.get("row"),
                    "rowHint": record.get("row"),
                    "modelHint": record.get("modelHint", ""),
                    "categoryHint": record.get("categoryHint", ""),
                    "partNumber": record.get("partNumber", ""),
                    "name": record.get("name", ""),
                    "partKey": record.get("partKey", ""),
                    "nameKey": record.get("nameKey", ""),
                },
            )
        return assets, "/wholesale-assets/wholesale-hero.png", placeholder

    image_files = sorted(
        [
            path
            for path in source_dir.iterdir()
            if path.suffix.lower() in {".png", ".jpg", ".jpeg", ".jp2"}
            and not path.name.startswith("EFIX_")
        ],
    )
    for index, src in enumerate(image_files, 1):
        suffix = ".png" if src.suffix.lower() == ".jp2" else src.suffix
        out_name = public_asset_name(index, suffix)
        out = public_dir / out_name
        width, height = copy_record_image(src, out)
        row_match = re.search(r"_G(\d+)_", src.name)
        assets.append(
            {
                "sourceName": src.name,
                "path": f"/wholesale-assets/{out_name}",
                "width": width,
                "height": height,
                "sheet": "",
                "sourceRow": None,
                "rowHint": int(row_match.group(1)) if row_match else None,
                "modelHint": "eSteer10"
                if "eSteer10" in src.name
                else ("eSteer20/20MAX" if "eSteer20" in src.name else ""),
                "categoryHint": "ブラケット" if "ブラケット" in src.name else "",
                "partNumber": "",
                "name": "",
                "partKey": "",
                "nameKey": "",
            },
        )
    return assets, "/wholesale-assets/wholesale-hero.png", placeholder


def is_model_compatible(item_model: str, asset_model: str) -> bool:
    if not asset_model:
        return True
    if asset_model == "共用":
        return True
    item_norm = item_model.lower().replace(" ", "")
    asset_norm = asset_model.lower().replace(" ", "")
    if "esteer10" in item_norm:
        return "esteer10" in asset_norm or asset_model == "共用"
    if "esteer20" in item_norm:
        return "esteer20" in asset_norm or asset_model == "共用"
    return True


def sheet_score(item: dict, asset: dict) -> int:
    sheet = asset.get("sheet", "")
    item_model = item.get("model", "").lower().replace(" ", "")
    category = item.get("category", "")
    score = 0

    if category == "ブラケット":
        score += 260 if "ブラケット" in sheet else -260
    elif "ブラケット" in sheet:
        score -= 180

    if "esteer10" in item_model or item_model == "esteer10":
        if "eSteer10" in sheet:
            score += 240
        elif "20" in sheet:
            score -= 260
    elif "esteer20" in item_model:
        if "20" in sheet:
            score += 240
        elif "eSteer10" in sheet:
            score -= 260

    return score


def score_asset(item: dict, asset: dict) -> int:
    score = sheet_score(item, asset)
    part_key = normalize_part(item.get("partNumber", ""))
    name_key = normalize_name(item.get("name", ""))
    short_key = normalize_name(item.get("shortName", ""))
    asset_part = asset.get("partKey", "")
    asset_name = asset.get("nameKey", "")

    if part_key and asset_part == part_key:
        score += 1000
    elif part_key and asset_part and (part_key in asset_part or asset_part in part_key):
        score += 320

    if name_key and asset_name == name_key:
        score += 520
    elif name_key and asset_name and (name_key in asset_name or asset_name in name_key):
        score += 220
    elif short_key and asset_name and (short_key in asset_name or asset_name in short_key):
        score += 80

    if is_model_compatible(item.get("model", ""), asset.get("modelHint", "")):
        score += 60
    else:
        score -= 200

    category = item.get("category", "")
    if category and category == asset.get("categoryHint", ""):
        score += 40

    return score


def pick_set_image(item: dict, assets: list[dict]) -> str | None:
    model_key = normalize_name(item.get("model", ""))
    if "20max" in model_key:
        matches = [
            asset
            for asset in assets
            if "esteer20max" in asset.get("nameKey", "") and "package" in asset.get("nameKey", "")
        ]
    elif "20" in model_key:
        matches = [
            asset
            for asset in assets
            if "esteer20" in asset.get("nameKey", "")
            and "20max" not in asset.get("nameKey", "")
            and "package" in asset.get("nameKey", "")
        ]
    elif "10" in model_key:
        matches = [asset for asset in assets if asset.get("nameKey") == normalize_name("eSteer 10セット")]
    else:
        matches = []
    return matches[0]["path"] if matches else None


def pick_image(item: dict, assets: list[dict], placeholder: str) -> str:
    if item.get("kind") == "set":
        set_image = pick_set_image(item, assets)
        if set_image:
            return set_image

    scored = sorted(
        ((score_asset(item, asset), asset) for asset in assets),
        key=lambda pair: pair[0],
        reverse=True,
    )
    if scored and scored[0][0] >= 450:
        return scored[0][1]["path"]
    return placeholder


def image_source(image_path: str, assets: list[dict]) -> dict:
    for asset in assets:
        if asset["path"] == image_path:
            return asset
    return {}


def workbook_sheet(wb, name: str):
    if name in wb.sheetnames:
        return wb[name]
    raise KeyError(f"sheet not found: {name}; available={wb.sheetnames}")


def read_catalog(
    price_path: Path,
    assets: list[dict],
    placeholder: str,
) -> tuple[list[CatalogItem], list[CatalogItem], list[dict]]:
    wb = load_workbook(price_path, read_only=True, data_only=True)
    diagnostics: list[dict] = []

    main_items: list[CatalogItem] = []
    main_ws = workbook_sheet(wb, "製品販売卸価格表")
    for excel_row, row in enumerate(main_ws.iter_rows(min_row=2, values_only=True), start=2):
        short_name = as_text(row[0])
        name = as_text(row[1])
        wholesale = as_int(row[2])
        retail = as_int(row[3])
        if not name or not wholesale or short_name.startswith("※"):
            continue
        model = name.replace("セット", "").strip()
        item = {
            "kind": "set",
            "shortName": short_name,
            "model": model,
            "category": "本体セット",
            "partNumber": "",
            "name": name,
        }
        image = pick_image(item, assets, placeholder)
        source = image_source(image, assets)
        diagnostics.append(
            {
                "id": f"set-{slug(short_name, str(excel_row))}",
                "name": name,
                "partNumber": "",
                "image": image,
                "imageSourceSheet": source.get("sheet", ""),
                "imageSourceRow": source.get("sourceRow"),
                "imageSourceName": source.get("sourceName", ""),
                "matched": image != placeholder,
            },
        )
        main_items.append(
            CatalogItem(
                id=f"set-{slug(short_name, str(excel_row))}",
                kind="set",
                shortName=short_name,
                model=model,
                section="本体",
                category="本体セット",
                partNumber="",
                name=name,
                requiredQty=1,
                wholesalePriceExTax=wholesale,
                retailPriceExTax=retail or None,
                wholesalePriceIncTax=price_inc_tax(wholesale) or 0,
                retailPriceIncTax=price_inc_tax(retail) if retail else None,
                image=image,
                sourceSheet=main_ws.title,
                sourceRow=excel_row,
            ),
        )

    option_items: list[CatalogItem] = []
    option_ws = workbook_sheet(wb, "オプション価格表")
    for excel_row, row in enumerate(option_ws.iter_rows(min_row=3, values_only=True), start=3):
        no = as_int(row[1])
        model = as_text(row[2])
        section = as_text(row[4])
        category = as_text(row[5])
        part_no = as_text(row[6])
        name = as_text(row[7])
        qty = as_int(row[8]) or 1
        wholesale = as_int(row[9])
        retail = as_int(row[10])
        short_name = as_text(row[0]) or name
        if not no or not name:
            continue
        item = {
            "kind": "part",
            "shortName": short_name,
            "model": model,
            "category": category,
            "partNumber": part_no,
            "name": name,
        }
        image = pick_image(item, assets, placeholder)
        item_id = f"part-{no:03d}-{slug(part_no or name, str(no))}"
        source = image_source(image, assets)
        diagnostics.append(
            {
                "id": item_id,
                "name": name,
                "partNumber": part_no,
                "image": image,
                "imageSourceSheet": source.get("sheet", ""),
                "imageSourceRow": source.get("sourceRow"),
                "imageSourceName": source.get("sourceName", ""),
                "matched": image != placeholder,
            },
        )
        option_items.append(
            CatalogItem(
                id=item_id,
                kind="part",
                shortName=short_name,
                model=model,
                section=section or "未分類",
                category=category or "その他",
                partNumber=part_no,
                name=name,
                requiredQty=qty,
                wholesalePriceExTax=wholesale,
                retailPriceExTax=retail or None,
                wholesalePriceIncTax=price_inc_tax(wholesale) or 0,
                retailPriceIncTax=price_inc_tax(retail) if retail else None,
                image=image,
                sourceSheet=option_ws.title,
                sourceRow=excel_row,
            ),
        )

    return main_items, option_items, diagnostics


def main() -> int:
    if len(sys.argv) != 6:
        print(
            "usage: generate_wholesale_catalog.py <price.xlsx> <image_dir> <generated_hero.png> <public_dir> <out_json>",
            file=sys.stderr,
        )
        return 2

    price_path = Path(sys.argv[1])
    image_dir = Path(sys.argv[2])
    generated_hero = Path(sys.argv[3])
    public_dir = Path(sys.argv[4])
    out_json = Path(sys.argv[5])

    assets, hero, placeholder = copy_images(image_dir, public_dir, generated_hero)
    main_items, option_items, diagnostics = read_catalog(price_path, assets, placeholder)

    payload = {
        "generatedAt": "2026-05-27",
        "taxRate": 0.1,
        "heroImage": hero,
        "mainItems": [asdict(item) for item in main_items],
        "optionItems": [asdict(item) for item in option_items],
        "imageAssets": assets,
        "imageDiagnostics": diagnostics,
    }
    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    matched = sum(1 for item in diagnostics if item["matched"])
    print(
        f"main={len(main_items)} options={len(option_items)} images={len(assets)} matched={matched}/{len(diagnostics)} out={out_json}",
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
