from __future__ import annotations

import csv
import hashlib
import json
import re
import sys
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw


def as_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_part(value: str) -> str:
    text = as_text(value).upper()
    if text.endswith(".0"):
        text = text[:-2]
    text = re.sub(r"[^A-Z0-9]+", "", text)
    return text.lstrip("0") if text.isdigit() else text


def normalize_name(value: str) -> str:
    text = as_text(value).lower()
    text = text.replace("ｅ", "e").replace("　", " ")
    return re.sub(r"[^a-z0-9ぁ-んァ-ヶ一-龥]+", "", text)


def file_slug(value: str, fallback: str) -> str:
    text = normalize_name(value)
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return (text or fallback)[:42]


def sheet_key(title: str) -> str:
    if "20" in title:
        return "esteer20"
    if "10" in title:
        return "esteer10"
    if "ブラケット" in title:
        return "bracket"
    return file_slug(title, "sheet")


def cell(ws, row: int, column: int) -> str:
    return as_text(ws.cell(row=row, column=column).value)


def row_info(ws, row: int, previous_names: dict[str, str]) -> dict[str, str]:
    title = ws.title
    if "ブラケット" in title:
        name = cell(ws, row, 1) or previous_names.get(title, "")
        part = cell(ws, row, 2)
        if cell(ws, row, 1):
            previous_names[title] = cell(ws, row, 1)
        model = "共用"
        text = f"{name} {part}".lower().replace(" ", "")
        if "esteer10のみ" in text or "esteer10用" in text:
            model = "eSteer10"
        elif "esteer20のみ" in text or "esteer20用" in text:
            model = "eSteer20/20MAX"
        return {
            "partNumber": part,
            "name": name,
            "modelHint": model,
            "categoryHint": "ブラケット",
        }

    part = cell(ws, row, 1)
    name = cell(ws, row, 2)
    model = "eSteer20/20MAX" if "20" in title else "eSteer10"
    category = "未分類"
    return {
        "partNumber": part,
        "name": name,
        "modelHint": model,
        "categoryHint": category,
    }


def save_png(raw: bytes, out_path: Path) -> tuple[int, int]:
    with Image.open(BytesIO(raw)) as img:
        converted = img.convert("RGBA") if img.mode in {"RGBA", "LA", "P"} else img.convert("RGB")
        converted.save(out_path, "PNG", optimize=True)
        return converted.size


def create_esteer10_set_image(out_dir: Path, records: list[dict]) -> dict | None:
    components = [
        record
        for key in {"J1AC01980900010002", "4090040041", "4006020041"}
        for record in records
        if record["partKey"] == key
    ]
    if len(components) < 3:
        return None

    canvas = Image.new("RGB", (960, 540), "white")
    draw = ImageDraw.Draw(canvas)
    x = 30
    for record in components[:3]:
        with Image.open(out_dir / record["fileName"]) as img:
            img = img.convert("RGBA")
            img.thumbnail((280, 360))
            y = 95 + (320 - img.height) // 2
            canvas.paste(img, (x + (280 - img.width) // 2, y), img)
            draw.rectangle((x, 430, x + 280, 431), fill=(0, 118, 83))
            x += 310
    out_name = "ExcelMapped_esteer10_set_components.png"
    canvas.save(out_dir / out_name, "PNG", optimize=True)
    return {
        "fileName": out_name,
        "sourceWorkbook": "generated",
        "sheet": "eSteer10",
        "row": 0,
        "column": 0,
        "partNumber": "",
        "name": "eSteer 10セット",
        "partKey": "",
        "nameKey": normalize_name("eSteer 10セット"),
        "modelHint": "eSteer10",
        "categoryHint": "本体セット",
        "width": 960,
        "height": 540,
        "hash": hashlib.sha1(out_name.encode("utf-8")).hexdigest()[:10],
    }


def extract(workbook_path: Path, out_dir: Path) -> list[dict]:
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(workbook_path, data_only=True)
    records: list[dict] = []
    previous_names: dict[str, str] = {}

    for ws in wb.worksheets:
        images = getattr(ws, "_images", [])
        if not images:
            continue
        for image in images:
            anchor = image.anchor._from
            row = anchor.row + 1
            column = anchor.col + 1
            info = row_info(ws, row, previous_names)
            raw = image._data()
            digest = hashlib.sha1(raw).hexdigest()[:10]
            part_key = normalize_part(info["partNumber"])
            name_key = normalize_name(info["name"])
            base = part_key or file_slug(info["name"], f"row{row}")
            out_name = f"ExcelMapped_{sheet_key(ws.title)}_R{row:03d}_{base}_{digest}.png"
            width, height = save_png(raw, out_dir / out_name)
            records.append(
                {
                    "fileName": out_name,
                    "sourceWorkbook": workbook_path.name,
                    "sheet": ws.title,
                    "row": row,
                    "column": column,
                    "partNumber": info["partNumber"],
                    "name": info["name"],
                    "partKey": part_key,
                    "nameKey": name_key,
                    "modelHint": info["modelHint"],
                    "categoryHint": info["categoryHint"],
                    "width": width,
                    "height": height,
                    "hash": digest,
                }
            )

    generated = create_esteer10_set_image(out_dir, records)
    if generated:
        records.append(generated)

    records.sort(key=lambda r: (r["modelHint"], r["categoryHint"], r["sheet"], r["row"], r["fileName"]))
    return records


def write_outputs(records: list[dict], out_dir: Path) -> None:
    json_path = out_dir / "EFIX_画像対応表.json"
    csv_path = out_dir / "EFIX_画像対応表.csv"
    json_path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")

    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "fileName",
                "sheet",
                "row",
                "partNumber",
                "name",
                "modelHint",
                "categoryHint",
                "width",
                "height",
            ],
        )
        writer.writeheader()
        for record in records:
            writer.writerow({key: record.get(key, "") for key in writer.fieldnames})


def main() -> int:
    if len(sys.argv) != 3:
        print("usage: extract_wholesale_image_assets.py <order.xlsx> <out_dir>", file=sys.stderr)
        return 2
    workbook_path = Path(sys.argv[1])
    out_dir = Path(sys.argv[2])
    records = extract(workbook_path, out_dir)
    write_outputs(records, out_dir)
    print(f"mapped_images={len(records)} out={out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
